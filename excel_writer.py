import json
import pandas as pd
import sys
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def json_to_excel(json_file, excel_file):
    try:
        # Load JSON and convert to DataFrame
        df = pd.read_json(json_file)
        df.to_excel(excel_file, index=False)

        # Load workbook and worksheet using openpyxl
        wb = load_workbook(excel_file)
        ws = wb.active
        
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue

        # Apply border to header row
        for cell in ws[1]:
            cell.fill = blue_fill

        # Find the index of the "Week" column (e.g., A, B, C...)
        week_col_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Week":
                week_col_index = idx
                break

        # Apply border to "Week" column (excluding header row which we already styled)
        if week_col_index:
            col_letter = get_column_letter(week_col_index)
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.fill = blue_fill

        # Set all column widths to 30
        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20

        # Save workbook
        wb.save(excel_file)

        print(f"✅ Styled Excel file created: {excel_file}")

    except Exception as e:
        print(f"❌ Failed to convert {json_file} to Excel: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 json_to_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    json_file = sys.argv[1]
    excel_file = sys.argv[2]

    json_to_excel(json_file, excel_file)
